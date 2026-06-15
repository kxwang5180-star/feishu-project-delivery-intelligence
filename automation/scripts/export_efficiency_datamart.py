from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean, median
from typing import Any, Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "data" / "efficiency_enhanced"
OUT_DIR = ROOT / "data" / "efficiency_datamart"
REPORT_START = "2025-01-01"
REPORT_END = "2026-05-31"


def parse_dt(value: Any) -> Optional[datetime]:
    if not value:
        return None
    text = str(value).strip().replace("Z", "+00:00")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:19] if "%H" in fmt else text[:10], fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def month_key(dt: datetime) -> str:
    return dt.strftime("%Y-%m")


def quarter_key(dt: datetime) -> str:
    return f"{dt.year}-Q{((dt.month - 1) // 3) + 1}"


def percentile(values: Iterable[Any], p: float) -> Optional[float]:
    nums = sorted(float(v) for v in values if v is not None and v != "")
    if not nums:
        return None
    if len(nums) == 1:
        return round(nums[0], 2)
    pos = (len(nums) - 1) * p
    lo = int(pos)
    hi = min(lo + 1, len(nums) - 1)
    return round(nums[lo] + (nums[hi] - nums[lo]) * (pos - lo), 2)


def avg(values: Iterable[Any]) -> Optional[float]:
    nums = [float(v) for v in values if v is not None and v != ""]
    return round(mean(nums), 2) if nums else None


def med(values: Iterable[Any]) -> Optional[float]:
    nums = [float(v) for v in values if v is not None and v != ""]
    return round(median(nums), 2) if nums else None


def load_rows() -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for line in (SOURCE_DIR / "enhanced_metrics.jsonl").read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        row = json.loads(line)
        online = parse_dt(row.get("online_time"))
        if not online:
            continue
        total = float(row.get("person_days") or 0)
        rows.append(
            {
                "work_item_id": row.get("work_item_id"),
                "demand_name": clean_text(row.get("name") or ""),
                "status": row.get("status") or "",
                "online_time": row.get("online_time"),
                "online_date": online.strftime("%Y-%m-%d"),
                "online_month": month_key(online),
                "online_quarter": quarter_key(online),
                "source_total_estimate_days": row.get("person_days"),
                "source_dev_estimate_days": row.get("dev_person_days"),
                "source_qa_estimate_days": row.get("qa_person_days"),
                "source_ui_estimate_days": row.get("ui_person_days"),
                "source_pm_estimate_days": row.get("pm_person_days"),
                "source_algo_estimate_days": row.get("algo_person_days"),
                "metric_delivery_effort_days": total,
                "metric_rd_days_excl_test": float(row.get("dev_person_days") or 0),
                "metric_test_days": float(row.get("qa_person_days") or 0),
                "new_size_group": "中小需求" if total <= 60 else "大需求",
                "legacy_size_band": row.get("size_band"),
                "legacy_calendar_delivery_cycle_days": row.get("delivery_cycle_days"),
                "earliest_schedule_start": row.get("earliest_schedule_start"),
                "participant_count": row.get("participant_count"),
                "dev_participant_count": row.get("dev_participant_count"),
                "is_report_month_included": REPORT_START[:7] <= month_key(online) <= REPORT_END[:7],
                "is_report_quarter_included": quarter_key(online) != "2026-Q2" and REPORT_START[:4] <= str(online.year) <= REPORT_END[:4],
                "source_cache_file": "data/efficiency_enhanced/enhanced_metrics.jsonl",
            }
        )
    return rows


def clean_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return " ".join(value.replace("\r", " ").replace("\n", " ").split())


def summarize(items: Sequence[Dict[str, Any]], key_name: Optional[str] = None, key_value: Optional[str] = None) -> Dict[str, Any]:
    result: Dict[str, Any] = {}
    if key_name:
        result[key_name] = key_value
    result.update(
        {
            "demand_count": len(items),
            "avg_delivery_effort_days": avg(item["metric_delivery_effort_days"] for item in items),
            "median_delivery_effort_days": med(item["metric_delivery_effort_days"] for item in items),
            "p25_delivery_effort_days": percentile((item["metric_delivery_effort_days"] for item in items), 0.25),
            "p75_delivery_effort_days": percentile((item["metric_delivery_effort_days"] for item in items), 0.75),
            "p90_delivery_effort_days": percentile((item["metric_delivery_effort_days"] for item in items), 0.90),
            "avg_rd_days": avg(item["metric_rd_days_excl_test"] for item in items),
            "median_rd_days": med(item["metric_rd_days_excl_test"] for item in items),
            "p90_rd_days": percentile((item["metric_rd_days_excl_test"] for item in items), 0.90),
            "avg_test_days": avg(item["metric_test_days"] for item in items),
            "median_test_days": med(item["metric_test_days"] for item in items),
            "p90_test_days": percentile((item["metric_test_days"] for item in items), 0.90),
        }
    )
    return result


def group_summary(rows: List[Dict[str, Any]], key: str, filter_key: Optional[str] = None) -> List[Dict[str, Any]]:
    usable = [row for row in rows if row.get(filter_key)] if filter_key else rows
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in usable:
        groups[str(row[key])].append(row)
    return [summarize(groups[name], key, name) for name in sorted(groups)]


def field_dictionary() -> List[Dict[str, Any]]:
    rows = [
        ("work_item_id", "需求ID", "work_item_id", "search_by_mql", "string", "飞书需求工作项ID", "主键"),
        ("demand_name", "需求名称", "name", "search_by_mql", "string", "需求标题", ""),
        ("status", "需求状态", "work_item_status", "search_by_mql", "string", "已结束/待验收/待推广", "当前分析过滤状态"),
        ("online_time", "项目上线时间", "field_584a64", "search_by_mql", "datetime", "项目上线时间", "月度/季度归属字段"),
        ("source_total_estimate_days", "投入总估分", "field_fba983", "search_by_mql", "number", "产品、前端、后端、测试等所有人员估分总和", "新口径交付周期来源"),
        ("source_dev_estimate_days", "开发估分", "field_db341e", "search_by_mql", "number", "研发人员估分总和", "研发时长来源，剔除测试"),
        ("source_ui_estimate_days", "UI估分", "field_040f76", "search_by_mql", "number", "UI/UX估分", "保留原字段"),
        ("source_qa_estimate_days", "测试估分", "field_715f2b", "search_by_mql", "number", "QA/测试人员估分总和", "测试时长来源"),
        ("source_pm_estimate_days", "产品估分", "field_142721", "search_by_mql", "number", "产品/PM估分", "保留原字段"),
        ("source_algo_estimate_days", "算法估分", "field_81ce0b", "search_by_mql", "number", "算法估分", "保留原字段"),
        ("metric_delivery_effort_days", "交付周期", "derived:field_fba983", "computed", "number", "所有人员估分总和", "新报告核心指标"),
        ("metric_rd_days_excl_test", "研发时长", "derived:field_db341e", "computed", "number", "研发人员估分总和，剔除测试", "新报告核心指标"),
        ("metric_test_days", "测试时长", "derived:field_715f2b", "computed", "number", "测试人员估分总和", "新报告核心指标"),
        ("new_size_group", "新需求分类", "derived:metric_delivery_effort_days", "computed", "string", "≤60为中小需求，>60为大需求", "新分类口径"),
        ("legacy_calendar_delivery_cycle_days", "旧自然日交付周期", "derived:get_node_detail+field_584a64", "computed", "number", "上线时间-最早排期开始时间", "历史报告口径，保留供追溯"),
        ("earliest_schedule_start", "最早排期开始时间", "derived:get_node_detail", "computed", "datetime", "节点/子任务最早排期开始时间", "历史报告口径"),
        ("participant_count", "参与人数", "derived:get_workitem_brief+get_node_detail", "computed", "number", "所有角色/节点/子任务人员去重", "历史报告指标"),
        ("dev_participant_count", "研发参与人数", "derived:role_members", "computed", "number", "QA/Server/FE/UI/架构/安全/合规角色去重", "历史报告指标"),
    ]
    return [
        {
            "column_name": col,
            "business_name": name,
            "source_field_id": fid,
            "source_method": method,
            "data_type": dtype,
            "definition": definition,
            "notes": notes,
        }
        for col, name, fid, method, dtype, definition, notes in rows
    ]


def metric_dictionary() -> List[Dict[str, Any]]:
    return [
        {"metric_id": "metric_delivery_effort_days", "metric_name": "交付周期", "formula": "field_fba983", "grain": "需求", "unit": "天/人天估分", "definition": "产品、前端、后端、测试等所有人员估分总和"},
        {"metric_id": "metric_rd_days_excl_test", "metric_name": "研发时长", "formula": "field_db341e", "grain": "需求", "unit": "天/人天估分", "definition": "研发人员估分总和，剔除测试"},
        {"metric_id": "metric_test_days", "metric_name": "测试时长", "formula": "field_715f2b", "grain": "需求", "unit": "天/人天估分", "definition": "测试人员估分总和"},
        {"metric_id": "new_size_group", "metric_name": "新需求分类", "formula": "IF(metric_delivery_effort_days<=60,'中小需求','大需求')", "grain": "需求", "unit": "分类", "definition": "按所有人员估分总和分类"},
        {"metric_id": "p90_*", "metric_name": "P90分位数", "formula": "percentile(metric,0.9)", "grain": "汇总周期/分类", "unit": "天/人天估分", "definition": "衡量长尾复杂需求风险"},
    ]


def connection_config() -> List[Dict[str, Any]]:
    return [
        {"connection_method": "Excel/PowerBI", "target": "efficiency_datamart.xlsx", "main_table": "demand_detail", "join_key": "work_item_id", "refresh_method": "重新运行 scripts/export_efficiency_datamart.py 后刷新文件"},
        {"connection_method": "CSV", "target": "demand_detail.csv", "main_table": "demand_detail", "join_key": "work_item_id", "refresh_method": "读取 data/efficiency_datamart/*.csv"},
        {"connection_method": "Python", "target": "manifest.json", "main_table": "demand_detail.csv", "join_key": "work_item_id", "refresh_method": "读取 manifest 中的 csv_paths"},
        {"connection_method": "数据库导入", "target": "demand_detail.csv", "main_table": "fact_demand_efficiency", "join_key": "work_item_id", "refresh_method": "按字段字典建表后导入CSV"},
    ]


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows([{key: clean_text(value) for key, value in row.items()} for row in rows])


def add_sheet(wb: Workbook, title: str, rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([clean_text(row.get(header)) for header in headers])
    header_fill = PatternFill("solid", fgColor="172033")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    table_ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    tab = Table(displayName=title.replace(" ", "_").replace("-", "_")[:30], ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = table_ref
    for idx, header in enumerate(headers, start=1):
        max_len = min(max(len(str(header)), *(len(str(row.get(header, ""))) for row in rows[:200])) + 2, 42)
        ws.column_dimensions[get_column_letter(idx)].width = max(10, max_len)


def write_workbook(path: Path, sheets: Dict[str, List[Dict[str, Any]]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for title, rows in sheets.items():
        add_sheet(wb, title, rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = load_rows()
    report_rows = [row for row in rows if row["online_month"] <= "2026-05"]
    quarter_rows = [row for row in report_rows if row["online_quarter"] != "2026-Q2"]
    monthly = group_summary(report_rows, "online_month")
    quarterly = group_summary(quarter_rows, "online_quarter")
    size_summary = group_summary(report_rows, "new_size_group")
    yoy = []
    for period, items in {
        "2025年1-5月": [row for row in report_rows if "2025-01" <= row["online_month"] <= "2025-05"],
        "2026年1-5月": [row for row in report_rows if "2026-01" <= row["online_month"] <= "2026-05"],
    }.items():
        yoy.append(summarize(items, "period", period))

    sheets = {
        "readme": [
            {"item": "dataset_name", "value": "department_demand_efficiency_datamart"},
            {"item": "generated_at", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
            {"item": "grain", "value": "one row per demand/work_item_id"},
            {"item": "primary_key", "value": "work_item_id"},
            {"item": "report_filter", "value": "monthly <= 2026-05; quarterly excludes 2026-Q2"},
            {"item": "source_cache", "value": "data/efficiency_enhanced/enhanced_metrics.jsonl"},
        ],
        "demand_detail": rows,
        "field_dictionary": field_dictionary(),
        "metric_dictionary": metric_dictionary(),
        "monthly_metrics": monthly,
        "quarterly_metrics": quarterly,
        "size_metrics": size_summary,
        "yoy_2026_1_5": yoy,
        "connection_config": connection_config(),
    }

    for name, data in sheets.items():
        write_csv(OUT_DIR / f"{name}.csv", data)
    workbook_path = OUT_DIR / "efficiency_datamart.xlsx"
    write_workbook(workbook_path, sheets)

    manifest = {
        "dataset": "department_demand_efficiency_datamart",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "primary_key": "work_item_id",
        "workbook": str(workbook_path),
        "csv_paths": {name: str(OUT_DIR / f"{name}.csv") for name in sheets},
        "source_cache": str(SOURCE_DIR / "enhanced_metrics.jsonl"),
        "field_dictionary": str(OUT_DIR / "field_dictionary.csv"),
        "metric_dictionary": str(OUT_DIR / "metric_dictionary.csv"),
    }
    (OUT_DIR / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "README.md").write_text(
        "\n".join(
            [
                "# 部门需求效率数据集市",
                "",
                "## 连接方式",
                "",
                "- Excel/PowerBI：连接 `efficiency_datamart.xlsx`，主表为 `demand_detail`，主键为 `work_item_id`。",
                "- CSV/数据库：导入 `demand_detail.csv`，字段定义见 `field_dictionary.csv`，指标定义见 `metric_dictionary.csv`。",
                "- Python/脚本：读取 `manifest.json` 中的 `csv_paths`。",
                "",
                "## 关键字段",
                "",
                "- `metric_delivery_effort_days`：交付周期，来源 `field_fba983`。",
                "- `metric_rd_days_excl_test`：研发时长，来源 `field_db341e`。",
                "- `metric_test_days`：测试时长，来源 `field_715f2b`。",
                "- `new_size_group`：`metric_delivery_effort_days <= 60` 为中小需求，否则为大需求。",
            ]
        ),
        encoding="utf-8",
    )
    print(workbook_path)
    print(OUT_DIR / "manifest.json")


if __name__ == "__main__":
    main()
