#!/usr/bin/env python3
"""Publish local CSV/XLSX tables into fixed Feishu online spreadsheet sheets."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any


ENV_FILES = [".env.local", ".env", "feishu.env", "feishu.env.md"]


def load_env(cwd: Path) -> dict[str, str]:
    values = dict(os.environ)
    for name in ENV_FILES:
        path = cwd / name
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            s = line.strip()
            if not s or s.startswith("#") or "=" not in s:
                continue
            key, value = s.split("=", 1)
            values.setdefault(key.strip(), value.strip().strip('"').strip("'"))
    return values


def request_json(method: str, url: str, token: str | None = None, body: Any | None = None, timeout: int = 30) -> dict[str, Any]:
    headers = {"Content-Type": "application/json; charset=utf-8"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code} calling {url}: {detail}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"Network error calling {url}: {exc}") from exc
    return json.loads(raw) if raw else {}


def tenant_access_token(base_url: str, app_id: str, app_secret: str) -> str:
    data = request_json(
        "POST",
        f"{base_url}/open-apis/auth/v3/tenant_access_token/internal",
        body={"app_id": app_id, "app_secret": app_secret},
    )
    if data.get("code") != 0:
        raise RuntimeError(f"Feishu auth failed: code={data.get('code')} msg={data.get('msg')}")
    token = data.get("tenant_access_token")
    if not token:
        raise RuntimeError("Feishu auth response did not include tenant_access_token")
    return token


def read_csv(path: Path) -> list[list[Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [row for row in csv.reader(f)]


def read_xlsx(path: Path, sheet_name: str | None = None) -> list[list[Any]]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise RuntimeError("Reading XLSX requires openpyxl. Use CSV or install openpyxl.") from exc
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        values = ["" if v is None else v for v in row]
        while values and values[-1] == "":
            values.pop()
        rows.append(values)
    while rows and not rows[-1]:
        rows.pop()
    return rows


def load_rows(source: Path, xlsx_sheet: str | None = None) -> list[list[Any]]:
    suffix = source.suffix.lower()
    if suffix == ".csv":
        return read_csv(source)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(source, xlsx_sheet)
    raise RuntimeError(f"Unsupported source type: {source}")


def col_letter(n: int) -> str:
    out = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def normalize_value(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return value


def normalize_rows(rows: list[list[Any]]) -> list[list[Any]]:
    width = max((len(r) for r in rows), default=0)
    return [[normalize_value(v) for v in r] + [""] * (width - len(r)) for r in rows]


def chunked(rows: list[list[Any]], size: int) -> list[list[list[Any]]]:
    return [rows[i : i + size] for i in range(0, len(rows), size)]


def write_ranges(base_url: str, token: str, spreadsheet_token: str, ranges: list[dict[str, Any]]) -> dict[str, Any]:
    body = {"valueRanges": ranges}
    data = request_json(
        "POST",
        f"{base_url}/open-apis/sheets/v2/spreadsheets/{urllib.parse.quote(spreadsheet_token)}/values_batch_update",
        token=token,
        body=body,
    )
    if data.get("code") != 0:
        raise RuntimeError(f"Feishu sheet write failed: code={data.get('code')} msg={data.get('msg')} data={data.get('data')}")
    return data


def spreadsheet_metainfo(base_url: str, token: str, spreadsheet_token: str) -> dict[str, Any]:
    data = request_json(
        "GET",
        f"{base_url}/open-apis/sheets/v2/spreadsheets/{urllib.parse.quote(spreadsheet_token)}/metainfo",
        token=token,
    )
    if data.get("code") != 0:
        raise RuntimeError(f"Feishu spreadsheet probe failed: code={data.get('code')} msg={data.get('msg')} data={data.get('data')}")
    return data


def blank_rows(row_count: int, col_count: int) -> list[list[str]]:
    return [[""] * col_count for _ in range(row_count)]


def resolve_path(path_text: str, base_dir: Path) -> Path:
    p = Path(path_text).expanduser()
    return p if p.is_absolute() else (base_dir / p).resolve()


def parse_token_from_url(url: str) -> tuple[str | None, str | None]:
    match = re.search(r"/sheets/([^/?#]+)", url)
    token = match.group(1) if match else None
    parsed = urllib.parse.urlparse(url)
    sheet_id = urllib.parse.parse_qs(parsed.query).get("sheet", [None])[0]
    return token, sheet_id


def main() -> int:
    parser = argparse.ArgumentParser(description="Publish local CSV/XLSX tables to a fixed Feishu spreadsheet.")
    parser.add_argument("--config", required=True, help="JSON config mapping sheet_id to source files")
    parser.add_argument("--dry-run", action="store_true", help="Validate config and print planned writes without calling Sheets write API")
    parser.add_argument("--probe-auth", action="store_true", help="Only test Feishu authentication")
    parser.add_argument("--probe-spreadsheet", action="store_true", help="Test access to the configured spreadsheet and print metadata")
    parser.add_argument("--chunk-rows", type=int, default=5000, help="Maximum rows per write request")
    args = parser.parse_args()

    cwd = Path.cwd()
    env = load_env(cwd)
    base_url = env.get("FEISHU_BASE_URL", "https://open.feishu.cn").rstrip("/")
    app_id = env.get("FEISHU_APP_ID")
    app_secret = env.get("FEISHU_APP_SECRET")
    if not app_id or not app_secret:
        raise SystemExit("Missing FEISHU_APP_ID or FEISHU_APP_SECRET in environment files")

    config_path = resolve_path(args.config, cwd)
    config = json.loads(config_path.read_text(encoding="utf-8"))
    config_dir = config_path.parent

    spreadsheet_token = config.get("spreadsheet_token") or env.get("FEISHU_SPREADSHEET_TOKEN")
    if not spreadsheet_token and config.get("spreadsheet_url"):
        spreadsheet_token, _ = parse_token_from_url(config["spreadsheet_url"])
    if not spreadsheet_token:
        raise SystemExit("Missing spreadsheet_token in config or FEISHU_SPREADSHEET_TOKEN")

    token = tenant_access_token(base_url, app_id, app_secret)
    if args.probe_auth:
        print(json.dumps({"ok": True, "auth": "ok", "base_url": base_url}, ensure_ascii=False))
        return 0
    if args.probe_spreadsheet:
        data = spreadsheet_metainfo(base_url, token, spreadsheet_token)
        print(json.dumps({"ok": True, "spreadsheet_token": spreadsheet_token, "metadata": data.get("data", data)}, ensure_ascii=False, indent=2))
        return 0

    clear_before_write = bool(config.get("clear_before_write", True))
    max_clear_rows = int(config.get("max_clear_rows", 2000))
    results: list[dict[str, Any]] = []

    for sheet in config.get("sheets", []):
        sheet_id = sheet.get("sheet_id")
        if not sheet_id and sheet.get("url"):
            _, sheet_id = parse_token_from_url(sheet["url"])
        if not sheet_id:
            raise SystemExit(f"Missing sheet_id for sheet config: {sheet}")
        source = resolve_path(sheet["source"], config_dir)
        rows = normalize_rows(load_rows(source, sheet.get("xlsx_sheet")))
        if not rows:
            rows = [[""]]
        cols = max((len(r) for r in rows), default=1)
        clear_cols = int(sheet.get("clear_cols", max(cols, int(config.get("min_clear_cols", cols)))))
        clear_rows = int(sheet.get("clear_rows", max(max_clear_rows, len(rows))))
        planned = {
            "name": sheet.get("name", sheet_id),
            "sheet_id": sheet_id,
            "source": str(source),
            "rows": len(rows),
            "cols": cols,
            "clear_before_write": clear_before_write,
        }
        if args.dry_run:
            results.append(planned)
            continue

        if clear_before_write:
            for offset, part in enumerate(chunked(blank_rows(clear_rows, clear_cols), args.chunk_rows)):
                start = offset * args.chunk_rows + 1
                end = start + len(part) - 1
                clear_range = f"{sheet_id}!A{start}:{col_letter(clear_cols)}{end}"
                write_ranges(base_url, token, spreadsheet_token, [{"range": clear_range, "values": part}])
                time.sleep(0.2)

        for offset, part in enumerate(chunked(rows, args.chunk_rows)):
            start = offset * args.chunk_rows + 1
            end = start + len(part) - 1
            data_range = f"{sheet_id}!A{start}:{col_letter(cols)}{end}"
            write_ranges(base_url, token, spreadsheet_token, [{"range": data_range, "values": part}])
            time.sleep(0.2)

        results.append({**planned, "status": "written"})

    print(json.dumps({"ok": True, "spreadsheet_token": spreadsheet_token, "results": results}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False, indent=2), file=sys.stderr)
        raise SystemExit(1)
